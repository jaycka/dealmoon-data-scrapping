from selenium import webdriver
import openpyxl
import time

browser = webdriver.Chrome()
browser.get('https://www.dealmoon.com/')
last_height = browser.execute_script("return document.body.scrollHeight")

while True:
    # Scroll down to bottom
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # Wait to load page
    time.sleep(1)

    # Calculate new scroll height and compare with last scroll height
    new_height = browser.execute_script("return document.body.scrollHeight")
    if new_height == last_height:
        break
    last_height = new_height
    
deals = browser.find_elements_by_css_selector('div.mlist.v2')
labels = browser.find_elements_by_css_selector('label.einfo')
links = browser.find_elements_by_class_name('right-cnt')
comments = browser.find_elements_by_css_selector('span.j-stat-count-comment.stat-count-comment')
fav = browser.find_elements_by_css_selector('span.j-stat-count-fav.event_statics_action')
share = browser.find_elements_by_css_selector('span.j-stat-count-share.stat-count-share')

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(len(deals)):
    try:
        sheet.cell(i+1,1).value = deals[i].get_attribute('data-dmt-d-value')
        print('deal '+str(i+1)+':'+deals[i].get_attribute('data-dmt-d-value'))
        
        sheet.cell(i+1,2).value = labels[i].find_element_by_css_selector('a.ib.ib-store.j-store').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+labels[i].find_element_by_css_selector('a.ib.ib-store.j-store').get_attribute('innerHTML'))
        
        sheet.cell(i+1,3).value = labels[i].find_element_by_css_selector('a.ib.ib-category').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+labels[i].find_element_by_css_selector('a.ib.ib-category').get_attribute('innerHTML'))
        
        sheet.cell(i+1,4).value = labels[i].find_element_by_css_selector('span.ib.published-date').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+labels[i].find_element_by_css_selector('span.ib.published-date').get_attribute('innerHTML'))

        sheet.cell(i+1,5).value = comments[i].find_element_by_css_selector('em.j-count').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+comments[i].find_element_by_css_selector('em.j-count').get_attribute('innerHTML'))

        sheet.cell(i+1,6).value = fav[i].find_element_by_class_name('j-count').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+fav[i].find_element_by_class_name('j-count').get_attribute('innerHTML'))

        sheet.cell(i+1,7).value = share[i].find_element_by_class_name('j-count').get_attribute('innerHTML')
        print('deal '+str(i+1)+':'+share[i].find_element_by_class_name('j-count').get_attribute('innerHTML'))

        sheet.cell(i+1,8).value = links[i].find_element_by_css_selector('a.zoom-title').get_attribute('href')
        print('deal '+str(i+1)+':'+links[i].find_element_by_css_selector('a.zoom-title').get_attribute('href'))
    except:
        pass

wb.save('dealmoon_deals.xlsx')
print('done')
